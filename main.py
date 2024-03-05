# -*- coding: utf-8 -*-
import datetime
import os
import configparser
import threading
import time
from openpyxl import Workbook, load_workbook
import shutil


def check_conditions(row):
    answer_time = float(row[answer_time_index].rstrip("min")) * 60  # 答题时间 紧跟最后一个题目后面
    answer_data = row[0:max_questions]  # 答题数据

    tmp_str = None
    # Rule 1: Answer time < 90 seconds
    if answer_time < min_answer_time:
        tmp_str = "答题时间过小"

    # Rule 2: Majority answer count / total questions > 90%; no Two reverse questions
    if max_same_answer > 0:  # 有最大答案相似度判断时
        if q_num_1st == 0:  # 没有反向题判断时
            answer_no_req = answer_data
        else:
            answer_no_req = answer_data[0: q_num_1st - 1] + answer_data[q_num_1st: max_questions - 1]

        majority_count = max(answer_no_req.count('一般'), answer_no_req.count('非常符合'),
                             answer_no_req.count('比较符合'),
                             answer_no_req.count('非常不符合'), answer_no_req.count('非常不符合'))
        if majority_count / max_questions > max_same_answer:
            if tmp_str is None:
                tmp_str = "同一答案相似度过大"
            else:
                tmp_str += ";同一答案相似度过大"

    # Rule 3: Two reverse questions have same or similar options
    if q_num_1st > 0:
        if (answer_data[q_num_1st - 1] == answer_data[q_num_2end - 1]
            and answer_data[q_num_1st - 1] != '一般') \
                or (answer_data[q_num_1st - 1] == '比较符合' and answer_data[q_num_2end - 1] == '非常符合') \
                or (answer_data[q_num_1st - 1] == '非常符合' and answer_data[q_num_2end - 1] == '比较符合') \
                or (answer_data[q_num_1st - 1] == '非常不符合' and answer_data[q_num_2end - 1] == '比较不符合') \
                or (answer_data[q_num_1st - 1] == '比较不符合' and answer_data[q_num_2end - 1] == '非常不符合'):
            if tmp_str is None:
                tmp_str = "反向题答案相似"
            else:
                tmp_str += ";反向题答案相似"

    return tmp_str


print_time_flag = True


def print_current_time():
    while print_time_flag:
        print("当前系统时间：", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        # 每隔10S打印一次
        time.sleep(10)


# 处理数据
def process_data(data, department_index):
    valid_data = {}
    invalid_data = {}
    for row in data:
        department = row[department_index]
        if department:
            departments = department.split('-')  # 数据格式为：视源股份-部件BG-TV事业部-营销中心-战队二部-销售
            # fo循环   3 个组织层级
            for i in range(3):
                output_folder = os.path.join('data_output', *departments[:2 + i])  # 从第二个字段开始计算
                valid_survey_file = os.path.join(output_folder, '有效问卷.xlsx')
                invalid_survey_file = os.path.join(output_folder, '无效问卷.xlsx')

                if not os.path.exists(output_folder):   # 创建新的目录
                    os.makedirs(output_folder)

                if valid_survey_file not in valid_data:     # 创建新的有效数据文件
                    valid_data[valid_survey_file] = []

                if invalid_survey_file not in invalid_data:     # 创建新的无效数据文件
                    invalid_data[invalid_survey_file] = []

                rule = check_conditions(row)    # 判断数据是否有效
                if rule:
                    invalid_data[invalid_survey_file].append(row + [rule])
                else:
                    valid_data[valid_survey_file].append(row)
    return valid_data, invalid_data


# 根据数据创建文件
def create_files_by_data(data_xlsx):
    # Write data to files
    index = 0
    all_index = len(data_xlsx)
    for valid_survey_file, rows in data_xlsx.items():
        if not os.path.exists(valid_survey_file):
            shutil.copyfile('template.xlsx', valid_survey_file)

        index += 1
        print(f' write valid_survey_file Processing {index}/{all_index}...')

        wb = load_workbook(valid_survey_file)
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(valid_survey_file)


if __name__ == '__main__':

    # 创建一个每隔10S打印当前时间的线程
    print_time_thread = threading.Thread(target=print_current_time)
    print_time_thread.start()

    # Read settings from config.ini
    encodings = ['utf-8', 'latin-1', 'gbk', 'gb2312', 'utf-16', 'utf-32', 'big5']  # Add more encodings if needed
    for encoding in encodings:
        try:
            config = configparser.ConfigParser()
            with open('config.ini', 'r', encoding=encoding) as f:
                config.read_file(f)

            min_answer_time = float(config.get('Settings', 'min_answer_time'))  # 最小答题时间
            q_num_1st = int(config.get('Settings', 'reverse_question_1'))  # 第1个相反题的题号
            q_num_2end = int(config.get('Settings', 'reverse_question_2'))  # 第2个相反题的题号
            max_questions = int(config.get('Settings', 'max_questions'))  # 最大题目数
            max_same_answer = float(config.get('Settings', 'max_same_answer'))  # 同一答案相似度阈值
            break
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            print("Config File not found.")
            exit()
        except Exception as e:
            print("Failed to open config file:", e)
            exit()

    # Read data from XLSX file
    data = []
    try:
        wb = load_workbook('data.xlsx')
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
    except FileNotFoundError:
        print("File not found.")
        exit()
    except Exception as e:
        print("Failed to open the file:", e)
        exit()

    # Find answer time index
    answer_time_index = None
    for i, label in enumerate(header):
        if label and '答题时间间隔' in label:
            answer_time_index = i
            break
    if answer_time_index is None:
        print("答题时间间隔列未找到")
        exit()
    # Find answer department index
    department_index = None
    for i, label in enumerate(header):
        if label and '答卷者部门' in label:
            department_index = i
            break
    if department_index is None:
        print("答题者所在部门未找到")
        exit()
    # Find 无效说明 index
    invalid_index = None
    for i, label in enumerate(header):
        if label and '无效说明' in label:
            invalid_index = i
            break
    # 表头添加无效数据字段
    if invalid_index is None:
        header.append("无效说明")
        invalid_index = len(header) - 1

    # Remove existing template.xlsx file if it exists
    if os.path.exists('template.xlsx'):
        os.remove('template.xlsx')
    # Create and Write head data to template.xlsx file
    wb_template = Workbook()
    ws_template = wb_template.active
    ws_template.append(header)
    wb_template.save('template.xlsx')

    # Remove existing data_output directory if it exists
    if os.path.exists('data_output'):
        shutil.rmtree('data_output')
    os.makedirs('data_output')
    # Process data and save to corresponding folders
    valid_data = {}
    invalid_data = {}
    valid_data, invalid_data = process_data(data, department_index)
    # Write data to files
    create_files_by_data(valid_data)
    create_files_by_data(invalid_data)
    # 关闭 print_time_thread 线程
    print_time_flag = False
    print_time_thread.join()
print("Data processed successfully.")
